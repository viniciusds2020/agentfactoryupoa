"""Parsing helpers for ingestion."""
from __future__ import annotations

import csv
import re
from pathlib import Path

from src.ingestion_cleaning import _PAGE_BREAK_MARKER, _PAGE_BREAK_RE, _normalize_structural_headers
from src.utils import get_logger, log_event

logger = get_logger(__name__)
_ARTIGO_RE = re.compile(
    r"^Art\.?\s*\d+(?:\s*(?:º|°|o|Âº|Â°|ª))?(?:\s*[-–—:]\s*.*)?$",
    re.IGNORECASE | re.MULTILINE,
)


def _build_pdf_page_context(blocks: list) -> dict[int, dict]:
    page_ctx: dict[int, dict] = {}
    for block in blocks or []:
        page = int(getattr(block, "page_number", 0) or 0)
        if page <= 0:
            continue
        entry = page_ctx.setdefault(page, {"block_types": {}, "section_hints": {}})
        btype = str(getattr(block, "block_type", "") or "").strip()
        if btype:
            bt = entry["block_types"]
            bt[btype] = int(bt.get(btype, 0)) + 1
        hint = str(getattr(block, "section_hint", "") or "").strip()
        if hint:
            sh = entry["section_hints"]
            sh[hint] = int(sh.get(hint, 0)) + 1

    compact: dict[int, dict] = {}
    for page, data in page_ctx.items():
        block_types = sorted(data["block_types"].items(), key=lambda x: (-x[1], x[0]))
        section_hints = sorted(data["section_hints"].items(), key=lambda x: (-x[1], x[0]))
        compact[page] = {
            "block_types": [k for k, _ in block_types[:4]],
            "section_hints": [k for k, _ in section_hints[:3]],
        }
    return compact


def _enrich_with_pdf_json_context(texts_for_embedding: list[str], metadatas: list[dict], pdf_blocks: list) -> list[str]:
    if not texts_for_embedding or not metadatas or not pdf_blocks:
        return texts_for_embedding
    page_ctx = _build_pdf_page_context(pdf_blocks)
    if not page_ctx:
        return texts_for_embedding

    enriched: list[str] = []
    for text, meta in zip(texts_for_embedding, metadatas):
        page = meta.get("page_number")
        if page is None:
            enriched.append(text)
            continue
        ctx = page_ctx.get(int(page))
        if not ctx:
            enriched.append(text)
            continue
        block_types = ctx.get("block_types", [])
        section_hints = ctx.get("section_hints", [])
        if block_types:
            meta["pdf_block_types"] = ",".join(block_types)
        if section_hints:
            meta["pdf_section_hints"] = " | ".join(section_hints)
        prefix_parts: list[str] = [f"[Pagina: {page}]"]
        if section_hints:
            prefix_parts.append(f"[Secoes: {'; '.join(section_hints)}]")
        if block_types:
            prefix_parts.append(f"[Tipos: {', '.join(block_types)}]")
        enriched.append(" ".join(prefix_parts) + f"\n{text}")
    return enriched


def _parse_xlsx(path: Path) -> tuple[str, dict]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    sheet_count = len(wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        lines: list[str] = [f"## Planilha: {sheet_name}\n"]
        header = rows[0]
        col_names = [str(cell) if cell is not None else "" for cell in header]
        lines.append("| " + " | ".join(col_names) + " |")
        lines.append("| " + " | ".join("---" for _ in col_names) + " |")
        for row in rows[1:]:
            cells = [str(cell) if cell is not None else "" for cell in row]
            if not any(c.strip() for c in cells):
                continue
            lines.append("| " + " | ".join(cells) + " |")
        parts.append("\n".join(lines))
    wb.close()
    return "\n\n".join(parts), {"sheet_count": sheet_count}


def _parse_csv(path: Path) -> tuple[str, dict]:
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file")

    if not rows:
        return "", {}

    lines: list[str] = []
    header = rows[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1:]:
        cells = row + [""] * (len(header) - len(row))
        cells = cells[: len(header)]
        if not any(c.strip() for c in cells):
            continue
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines), {"row_count": len(rows) - 1}


def _parse_docling_fast(path: Path) -> tuple[str, dict] | None:
    try:
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.document_converter import DocumentConverter, PdfFormatOption

        pdf_opts = PdfPipelineOptions()
        pdf_opts.do_ocr = False
        pdf_opts.do_table_structure = True
        converter = DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts)})
        result = converter.convert(str(path))
        text = result.document.export_to_markdown(page_break_placeholder="<!-- PAGE_BREAK -->")
        if len(text.strip()) < 200:
            return None
        page_count = len(result.pages) if hasattr(result, "pages") else None
        meta: dict = {"parser": "docling"}
        if page_count is not None:
            meta["page_count"] = page_count
        return text, meta
    except Exception as exc:
        log_event(logger, 30, "docling fast parse failed", path=str(path), error=str(exc))
        return None


def _has_line_structure(text: str) -> bool:
    normalized = _normalize_structural_headers(text)
    art_count = len(_ARTIGO_RE.findall(normalized))
    section_count = len(re.findall(r"^CAP[ÍI]TULO\s", text, re.IGNORECASE | re.MULTILINE))
    return (art_count + section_count) >= 3


def _parse_pdf(path: Path) -> tuple[str, dict]:
    docling_result = _parse_docling_fast(path)
    if docling_result is not None:
        text, meta = docling_result
        if _has_line_structure(text):
            log_event(logger, 20, "PDF parsed with docling (markdown with headings)", path=str(path))
            return docling_result
        log_event(logger, 20, "Docling output lacks line structure, trying PyMuPDF", path=str(path))

    try:
        import fitz
    except ImportError:
        if docling_result is not None:
            return docling_result
        log_event(logger, 30, "PyMuPDF not available, falling back to docling+OCR", path=str(path))
        return _parse_docling(path)

    doc = fitz.open(str(path))
    page_count = len(doc)
    pages: list[str] = []
    for page in doc:
        text = page.get_text("text")
        if text.strip():
            pages.append(text.strip())
    doc.close()

    total_chars = sum(len(p) for p in pages)
    if total_chars > 200:
        log_event(logger, 20, "PDF parsed with PyMuPDF fallback (plain text)", path=str(path))
        return _PAGE_BREAK_MARKER.join(pages), {"page_count": page_count, "parser": "pymupdf"}

    log_event(logger, 20, "Both docling-fast and PyMuPDF extracted too little, using docling+OCR", path=str(path), pymupdf_chars=total_chars)
    return _parse_docling(path)


def _parse_docling(path: Path) -> tuple[str, dict]:
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pdf_opts = PdfPipelineOptions()
    pdf_opts.do_ocr = True
    pdf_opts.do_table_structure = True
    converter = DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts)})
    result = converter.convert(str(path))
    text = result.document.export_to_markdown(page_break_placeholder="<!-- PAGE_BREAK -->")
    page_count = len(result.pages) if hasattr(result, "pages") else None
    meta: dict = {"parser": "docling"}
    if page_count is not None:
        meta["page_count"] = page_count
    return text, meta


def _parse(source: str) -> tuple[str, dict]:
    path = Path(source)
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                return path.read_text(encoding=encoding), {}
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Unable to decode text file: {suffix}")
    if suffix in {".xlsx", ".xls"}:
        return _parse_xlsx(path)
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix == ".pdf":
        return _parse_pdf(path)
    if suffix in {".docx", ".pptx"}:
        return _parse_docling(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def _assign_page_numbers(chunks: list[str], raw_text: str) -> list[int | None]:
    marker_positions: list[int] = [m.start() for m in _PAGE_BREAK_RE.finditer(raw_text)]
    if not marker_positions:
        return [None] * len(chunks)
    clean_text = _PAGE_BREAK_RE.sub("", raw_text)
    page_boundaries: list[int] = []
    offset_adjustment = 0
    for pos in marker_positions:
        clean_pos = pos - offset_adjustment
        page_boundaries.append(clean_pos)
        offset_adjustment += len("<!-- PAGE_BREAK -->")

    def _page_for_position(pos: int) -> int:
        page = 1
        for boundary in page_boundaries:
            if pos >= boundary:
                page += 1
            else:
                break
        return page

    result: list[int | None] = []
    for chunk in chunks:
        idx = clean_text.find(chunk[:80])
        result.append(_page_for_position(idx) if idx >= 0 else None)
    return result
