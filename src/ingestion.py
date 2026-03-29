"""Document ingestion pipeline: parse -> chunk -> embed -> upsert to FAISS."""
from __future__ import annotations

import os
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from src.utils import chunk_id, get_logger, log_event

logger = get_logger(__name__)

# Abreviações PT-BR comuns que não devem ser tratadas como fim de frase
_ABBREVIATIONS = re.compile(
    r"\b(Art|Inc|Sr|Sra|Dr|Dra|Prof|Ltda|S\.A|Jr|n|p|fl|fls|vol|ed|etc|ex)\.",
    re.IGNORECASE,
)
_ABBREV_PLACEHOLDER = "\x00"

# Marker inserted by docling between pages when page_break_placeholder is set.
_PAGE_BREAK_MARKER = "\n<!-- PAGE_BREAK -->\n"
_PAGE_BREAK_RE = re.compile(r"<!-- PAGE_BREAK -->")

# Patterns for repeated page headers/footers injected by docling.
# These add noise to chunks and hurt semantic search.
_PAGE_HEADER_PATTERNS = [
    # Generic doc management headers (e.g., "Estatuto Social Nº/Rev.: ... Este documento faz parte ...")
    re.compile(
        r"^[^\n]{0,200}(?:Nº/Rev\.?:|N°/Rev\.?:)[^\n]*Este documento faz parte[^\n]*\n?",
        re.MULTILINE | re.IGNORECASE,
    ),
    # Download watermarks (e.g., "(Baixado por FULANO em DD/MM/YYYY HH:MM:SS")
    re.compile(
        r"\(Baixado por [^\n)]+\)\s*",
        re.IGNORECASE,
    ),
    # "Página X de Y" footers
    re.compile(r"Página\s+\d+\s+de\s+\d+\s*", re.IGNORECASE),
    # "Classificação da Informação: ..." lines
    re.compile(r"Classificação da Informação:\s*\w+\s*", re.IGNORECASE),
    # Standalone image placeholders
    re.compile(r"<!--\s*image\s*-->\s*", re.IGNORECASE),
]


def _strip_page_headers(text: str) -> str:
    """Remove repeated page headers, footers, and watermarks from parsed text."""
    for pattern in _PAGE_HEADER_PATTERNS:
        text = pattern.sub("", text)
    # Collapse multiple blank lines left behind
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ── Heavy structural cleanup (Gap 6) ────────────────────────────────────────

_HYPHENATED_BREAK_RE = re.compile(r"(\w)-\s*\n\s*(\w)")
_BROKEN_LINE_RE = re.compile(r"(?<=[a-záàâãéèêíóôõúüç,;])\s*\n\s*(?=[a-záàâãéèêíóôõúüç])", re.IGNORECASE)
_REPEATED_BLOCK_MIN_LEN = 40  # minimum length to consider a block for dedup


def _merge_broken_lines(text: str) -> str:
    """Repair lines broken mid-word (hyphenation) or mid-sentence.

    Joins 'respons-\\nável' → 'responsável' and lines ending in lowercase
    followed by lowercase continuation (likely broken by PDF extraction).
    """
    # First: merge hyphenated breaks (word-\nword → wordword)
    text = _HYPHENATED_BREAK_RE.sub(r"\1\2", text)
    # Second: merge lines broken mid-sentence (lowercase,\nlowercase)
    text = _BROKEN_LINE_RE.sub(" ", text)
    return text


def _deduplicate_paragraphs(text: str) -> str:
    """Remove duplicate paragraphs that appear from repeated headers/sections.

    Keeps the first occurrence of each paragraph. Only deduplicates blocks
    longer than _REPEATED_BLOCK_MIN_LEN to avoid removing legitimate short lines.
    """
    lines = text.split("\n\n")
    seen: set[str] = set()
    result: list[str] = []
    for block in lines:
        stripped = block.strip()
        if len(stripped) < _REPEATED_BLOCK_MIN_LEN:
            result.append(block)
            continue
        # Normalize for comparison (collapse whitespace, lowercase)
        key = re.sub(r"\s+", " ", stripped.lower())
        if key in seen:
            continue
        seen.add(key)
        result.append(block)
    return "\n\n".join(result)


def _remove_orphan_page_numbers(text: str) -> str:
    """Remove standalone page numbers (lines containing only a number)."""
    lines = text.splitlines()
    result = [line for line in lines if not re.fullmatch(r"\s*\d{1,4}\s*", line)]
    return "\n".join(result)


def _normalize_whitespace(text: str) -> str:
    """Normalize excessive whitespace while preserving paragraph breaks."""
    # Collapse 3+ blank lines to 2
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Collapse multiple spaces to single (but not newlines)
    text = re.sub(r"[^\S\n]+", " ", text)
    return text.strip()


def deep_clean_text(text: str) -> str:
    """Apply heavy structural cleanup for reliable downstream processing.

    Pipeline:
    1. Merge broken lines (hyphenation and mid-sentence breaks)
    2. Remove orphan page numbers
    3. Deduplicate repeated paragraphs
    4. Normalize whitespace
    """
    text = _merge_broken_lines(text)
    text = _remove_orphan_page_numbers(text)
    text = _deduplicate_paragraphs(text)
    text = _normalize_whitespace(text)
    return text


_MARKDOWN_HEADING_RE = re.compile(r"^#{1,6}\s+", re.MULTILINE)
_WRAPPED_STRUCTURAL_LINE_RE = re.compile(
    r"^\s*[\[(]\s*((?:T[IÃ]TULO|CAP[IÃ]TULO|SE[CÃ‡][AÃƒ]O)\s+[IVXLCDM\d]+(?:\s*[-â€“â€”:\.]\s*.+)?)\s*[\])]\s*$",
    re.IGNORECASE | re.MULTILINE,
)


def _normalize_markdown(text: str) -> str:
    """Strip markdown heading prefixes so section detection regexes work.

    Converts ``## CAPÍTULO I - Foo`` to ``CAPÍTULO I - Foo``.
    Preserves the heading text and line position.
    """
    return _MARKDOWN_HEADING_RE.sub("", text)


def _normalize_structural_headers(text: str) -> str:
    """Normalize wrapped headings so section regexes can detect them."""
    normalized_lines: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if (
            stripped.startswith(("[", "("))
            and stripped.endswith(("]", ")"))
            and any(token in stripped.upper() for token in ("CAP", "TIT", "SE"))
        ):
            normalized_lines.append(stripped[1:-1].strip())
        else:
            normalized_lines.append(line)
    return "\n".join(normalized_lines)


# ── Section/chapter detection ────────────────────────────────────────────────

_SECTION_LEVEL_PATTERNS: list[tuple[int, str, re.Pattern]] = [
    # level, label key, pattern
    (1, "title", re.compile(
        r"^(T[IÍ]TULO\s+[IVXLCDM\d]+(?:\s*[-–—:\.]\s*.+)?)\s*$",
        re.IGNORECASE | re.MULTILINE,
    )),
    (2, "section", re.compile(
        r"^(CAP[IÍ]TULO\s+[IVXLCDM\d]+(?:\s*[-–—:\.]\s*.+)?)\s*$",
        re.IGNORECASE | re.MULTILINE,
    )),
    (3, "subsection", re.compile(
        r"^(SE[CÇ][AÃ]O\s+[IVXLCDM\d]+(?:\s*[-–—:\.]\s*.+)?)\s*$",
        re.IGNORECASE | re.MULTILINE,
    )),
]


class _SectionMarker:
    __slots__ = ("offset", "header", "level", "key")

    def __init__(self, offset: int, header: str, level: int, key: str):
        self.offset = offset
        self.header = header.strip()
        self.level = level
        self.key = key


def _normalize_heading_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized.upper().strip()


def _line_level_key(line: str) -> tuple[int, str] | None:
    norm = _normalize_heading_text(line)
    if re.match(r"^TITULO\s+([IVXLCDM]+|\d{1,4})\b", norm):
        return (1, "title")
    if re.match(r"^CAPITULO\s+([IVXLCDM]+|\d{1,4})\b", norm):
        return (2, "section")
    if re.match(r"^SECAO\s+([IVXLCDM]+|\d{1,4})\b", norm):
        return (3, "subsection")
    return None


def _detect_sections(text: str) -> list[_SectionMarker]:
    """Scan text for structural headers (Título, Capítulo, Seção) and return ordered markers."""
    markers: list[_SectionMarker] = []
    for level, key, pattern in _SECTION_LEVEL_PATTERNS:
        for m in pattern.finditer(text):
            markers.append(_SectionMarker(offset=m.start(), header=m.group(1).strip(), level=level, key=key))

    # Fallback robusto: detecção por linha com normalização (acento-insensível),
    # cobre extrações OCR/pipelines onde regex acentuada falha.
    if not markers or not any(m.key == "section" for m in markers):
        offset = 0
        for line in text.splitlines(keepends=True):
            raw_line = line.rstrip("\r\n")
            lk = _line_level_key(raw_line)
            if lk:
                level, key = lk
                markers.append(_SectionMarker(offset=offset, header=raw_line.strip(), level=level, key=key))
            offset += len(line)

    markers.sort(key=lambda m: m.offset)
    return markers


def _assign_sections(chunks: list[str], markers: list[_SectionMarker], clean_text: str) -> list[dict]:
    """Map each chunk to its enclosing section(s) based on position in the text.

    Returns a list of dicts with keys like ``title``, ``section``, ``subsection``
    populated when a chunk falls after a detected header.
    """
    if not markers:
        return [{} for _ in chunks]

    result: list[dict] = []
    for chunk in chunks:
        idx = clean_text.find(chunk[:80])
        info: dict[str, str] = {}
        if idx >= 0:
            # For each level, find the most recent marker before this chunk
            for level, key, _ in _SECTION_LEVEL_PATTERNS:
                best: _SectionMarker | None = None
                for m in markers:
                    if m.level == level and m.offset <= idx:
                        best = m
                    elif m.offset > idx:
                        break
                if best:
                    info[best.key] = best.header
        result.append(info)
    return result


def _prepend_section_context(chunks: list[str], section_info: list[dict]) -> list[str]:
    """Prepend a section breadcrumb to each chunk so embeddings capture structural context."""
    result: list[str] = []
    for chunk, info in zip(chunks, section_info):
        if info:
            breadcrumb = " > ".join(info.values())
            result.append(f"[{breadcrumb}]\n{chunk}")
        else:
            result.append(chunk)
    return result


# ── Legal document structural chunking ───────────────────────────────────────

_ARTIGO_RE = re.compile(
    r"^(Art\.?\s*\d+[°ºª]?\s*[-–—.:]\s*)",
    re.IGNORECASE | re.MULTILINE,
)

_PARAGRAFO_RE = re.compile(
    r"^((?:§\s*\d+[°ºª]?\s*[-–—.:]\s*)|(?:Parágrafo\s+[Úú]nico\s*[-–—.:]\s*))",
    re.IGNORECASE | re.MULTILINE,
)

_INCISO_RE = re.compile(
    r"^((?:[IVXLCDM]+|\d+)\s*[-–—]\s+)",
    re.IGNORECASE | re.MULTILINE,
)

_ARTIGO_LABEL_RE = re.compile(r"Art\.?\s*(\d+)", re.IGNORECASE)


def _is_legal_document(text: str, threshold: int = 3) -> bool:
    """Auto-detect whether parsed text has legal document structure."""
    sample = text[:5000]
    count = len(_ARTIGO_RE.findall(sample))
    count += len(re.findall(r"CAP[IÍ]TULO|SE[CÇ][AÃ]O|T[IÍ]TULO", sample, re.IGNORECASE))
    return count >= threshold


@dataclass
class LegalChunk:
    text: str
    chunk_type: str  # "parent" | "child" | "general"
    artigo: str = ""
    paragrafo: str = ""
    inciso: str = ""
    capitulo: str = ""
    secao: str = ""
    titulo: str = ""
    caminho_hierarquico: str = ""
    parent_key: str = ""
    references: str = ""


_INTERNAL_REF_RE = re.compile(r"art(?:igo)?\.?\s*(\d+)", re.IGNORECASE)


def _extract_references(text: str, own_artigo_num: str = "") -> str:
    """Extract internal article references from chunk text, excluding self-references."""
    refs = set()
    for m in _INTERNAL_REF_RE.finditer(text):
        num = m.group(1)
        if num != own_artigo_num:
            refs.add(f"Art. {num}")
    return ", ".join(sorted(refs)) if refs else ""


def _section_for_offset(markers: list[_SectionMarker], offset: int) -> dict[str, str]:
    """Find the enclosing sections for a given character offset."""
    info: dict[str, str] = {}
    for level, key, _ in _SECTION_LEVEL_PATTERNS:
        best: _SectionMarker | None = None
        for m in markers:
            if m.level == level and m.offset <= offset:
                best = m
            elif m.offset > offset:
                break
        if best:
            info[best.key] = best.header
    return info


def _split_legal(
    text: str,
    section_markers: list[_SectionMarker],
    doc_id: str,
    max_chunk_size: int = 2000,
    child_threshold: int = 800,
) -> list[LegalChunk]:
    """Split legal text by article structure into parent and child chunks.

    - Parent chunk = full article (caput + all §§ + incisos)
    - Child chunks = individual §§ when article exceeds *child_threshold*
    - Preamble text before first article uses generic chunking
    """
    # Find all article boundaries
    article_matches = list(_ARTIGO_RE.finditer(text))

    if not article_matches:
        # No articles found — treat as generic document
        return [LegalChunk(text=text, chunk_type="general")]

    chunks: list[LegalChunk] = []

    # Handle preamble (text before first article)
    preamble = text[: article_matches[0].start()].strip()
    if preamble:
        preamble_info = _section_for_offset(section_markers, 0)
        path = " > ".join(preamble_info.values()) if preamble_info else ""
        chunks.append(LegalChunk(
            text=preamble,
            chunk_type="general",
            caminho_hierarquico=path,
            **{k: v for k, v in preamble_info.items() if k in ("titulo", "capitulo", "secao")},
        ))

    # Process each article
    for i, match in enumerate(article_matches):
        art_start = match.start()
        art_end = article_matches[i + 1].start() if i + 1 < len(article_matches) else len(text)
        art_text = text[art_start:art_end].strip()

        if not art_text:
            continue

        # Extract article number
        label_match = _ARTIGO_LABEL_RE.search(art_text)
        art_num = label_match.group(1) if label_match else str(i + 1)
        art_label = f"Art. {art_num}"

        # Find enclosing sections
        sec_info = _section_for_offset(section_markers, art_start)

        # Build hierarchical path
        path_parts = list(sec_info.values())
        path_parts.append(art_label)
        caminho = " > ".join(path_parts)

        parent_key = f"{doc_id}::{art_label}"
        refs = _extract_references(art_text, art_num)

        # Always create parent chunk with the full article text so structural
        # citations and parent expansion keep the legal unit intact.
        parent_text = art_text
        chunks.append(LegalChunk(
            text=parent_text,
            chunk_type="parent",
            artigo=art_label,
            capitulo=sec_info.get("section", ""),
            secao=sec_info.get("subsection", ""),
            titulo=sec_info.get("title", ""),
            caminho_hierarquico=caminho,
            parent_key=parent_key,
            references=refs,
        ))

        # Create child chunks for long articles
        if len(art_text) > child_threshold:
            para_matches = list(_PARAGRAFO_RE.finditer(art_text))
            inciso_matches = list(_INCISO_RE.finditer(art_text))
            if para_matches:
                # Caput = text before first paragraph
                caput_text = art_text[: para_matches[0].start()].strip()
                if caput_text:
                    chunks.append(LegalChunk(
                        text=caput_text,
                        chunk_type="child",
                        artigo=art_label,
                        paragrafo="caput",
                        inciso="",
                        capitulo=sec_info.get("section", ""),
                        secao=sec_info.get("subsection", ""),
                        titulo=sec_info.get("title", ""),
                        caminho_hierarquico=f"{caminho} > caput",
                        parent_key=parent_key,
                        references=_extract_references(caput_text, art_num),
                    ))

                # Each paragraph as child
                for j, pm in enumerate(para_matches):
                    p_start = pm.start()
                    p_end = para_matches[j + 1].start() if j + 1 < len(para_matches) else len(art_text)
                    p_text = art_text[p_start:p_end].strip()
                    if not p_text:
                        continue

                    # Extract paragraph label
                    p_header = pm.group(1).strip().rstrip("-–—.: ")
                    chunks.append(LegalChunk(
                        text=p_text,
                        chunk_type="child",
                        artigo=art_label,
                        paragrafo=p_header,
                        inciso="",
                        capitulo=sec_info.get("section", ""),
                        secao=sec_info.get("subsection", ""),
                        titulo=sec_info.get("title", ""),
                        caminho_hierarquico=f"{caminho} > {p_header}",
                        parent_key=parent_key,
                        references=_extract_references(p_text, art_num),
                    ))
            elif inciso_matches:
                caput_text = art_text[: inciso_matches[0].start()].strip()
                if caput_text:
                    chunks.append(LegalChunk(
                        text=caput_text,
                        chunk_type="child",
                        artigo=art_label,
                        paragrafo="caput",
                        inciso="",
                        capitulo=sec_info.get("section", ""),
                        secao=sec_info.get("subsection", ""),
                        titulo=sec_info.get("title", ""),
                        caminho_hierarquico=f"{caminho} > caput",
                        parent_key=parent_key,
                        references=_extract_references(caput_text, art_num),
                    ))

                for j, im in enumerate(inciso_matches):
                    i_start = im.start()
                    i_end = inciso_matches[j + 1].start() if j + 1 < len(inciso_matches) else len(art_text)
                    i_text = art_text[i_start:i_end].strip()
                    if not i_text:
                        continue

                    inciso_label = im.group(1).strip().rstrip("-–— ")
                    chunks.append(LegalChunk(
                        text=i_text,
                        chunk_type="child",
                        artigo=art_label,
                        paragrafo="",
                        inciso=inciso_label,
                        capitulo=sec_info.get("section", ""),
                        secao=sec_info.get("subsection", ""),
                        titulo=sec_info.get("title", ""),
                        caminho_hierarquico=f"{caminho} > inciso {inciso_label}",
                        parent_key=parent_key,
                        references=_extract_references(i_text, art_num),
                    ))

    return chunks


def _build_contextualized_text(chunk: LegalChunk, doc_id: str = "") -> str:
    """Build enriched text for embedding (not for storage)."""
    parts: list[str] = []
    if doc_id:
        parts.append(f"[Documento: {doc_id}]")
    if chunk.caminho_hierarquico:
        parts.append(f"[Caminho: {chunk.caminho_hierarquico}]")
    if chunk.inciso:
        parts.append(f"[Inciso: {chunk.inciso}]")
    parts.append(chunk.text)
    return "\n".join(parts)


def _build_pdf_page_context(blocks: list) -> dict[int, dict]:
    """Build compact JSON-like page context from pdf_pipeline blocks."""
    page_ctx: dict[int, dict] = {}
    for block in blocks or []:
        page = int(getattr(block, "page_number", 0) or 0)
        if page <= 0:
            continue
        entry = page_ctx.setdefault(page, {"block_types": {}, "section_hints": {}})

        btype = str(getattr(block, "block_type", "") or "").strip()
        if btype:
            bt = entry["block_types"]
            bt[btype] = int(bt.get(btype, 0)) + 1

        hint = str(getattr(block, "section_hint", "") or "").strip()
        if hint:
            sh = entry["section_hints"]
            sh[hint] = int(sh.get(hint, 0)) + 1

    compact: dict[int, dict] = {}
    for page, data in page_ctx.items():
        block_types = sorted(
            data["block_types"].items(),
            key=lambda x: (-x[1], x[0]),
        )
        section_hints = sorted(
            data["section_hints"].items(),
            key=lambda x: (-x[1], x[0]),
        )
        compact[page] = {
            "block_types": [k for k, _ in block_types[:4]],
            "section_hints": [k for k, _ in section_hints[:3]],
        }
    return compact


def _enrich_with_pdf_json_context(
    texts_for_embedding: list[str],
    metadatas: list[dict],
    pdf_blocks: list,
) -> list[str]:
    """Enrich embedding text and metadata with compact JSON structural context."""
    if not texts_for_embedding or not metadatas or not pdf_blocks:
        return texts_for_embedding

    page_ctx = _build_pdf_page_context(pdf_blocks)
    if not page_ctx:
        return texts_for_embedding

    enriched: list[str] = []
    for text, meta in zip(texts_for_embedding, metadatas):
        page = meta.get("page_number")
        if page is None:
            enriched.append(text)
            continue

        ctx = page_ctx.get(int(page))
        if not ctx:
            enriched.append(text)
            continue

        block_types = ctx.get("block_types", [])
        section_hints = ctx.get("section_hints", [])

        # Persist compact context for filtering/reranking.
        if block_types:
            meta["pdf_block_types"] = ",".join(block_types)
        if section_hints:
            meta["pdf_section_hints"] = " | ".join(section_hints)

        prefix_parts: list[str] = [f"[Pagina: {page}]"]
        if section_hints:
            prefix_parts.append(f"[Secoes: {'; '.join(section_hints)}]")
        if block_types:
            prefix_parts.append(f"[Tipos: {', '.join(block_types)}]")
        prefix = " ".join(prefix_parts)
        enriched.append(f"{prefix}\n{text}")

    return enriched


def _parse_xlsx(path: Path) -> tuple[str, dict]:
    """Parse XLSX/XLS into markdown-style text preserving table structure."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    sheet_count = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build markdown table per sheet
        lines: list[str] = [f"## Planilha: {sheet_name}\n"]
        header = rows[0]
        col_names = [str(cell) if cell is not None else "" for cell in header]

        # Markdown table header
        lines.append("| " + " | ".join(col_names) + " |")
        lines.append("| " + " | ".join("---" for _ in col_names) + " |")

        # Data rows
        for row in rows[1:]:
            cells = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if not any(c.strip() for c in cells):
                continue
            lines.append("| " + " | ".join(cells) + " |")

        parts.append("\n".join(lines))

    wb.close()
    text = "\n\n".join(parts)
    return text, {"sheet_count": sheet_count}


def _parse_csv(path: Path) -> tuple[str, dict]:
    """Parse CSV into markdown-style text preserving table structure."""
    import csv

    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file")

    if not rows:
        return "", {}

    # Build markdown table
    lines: list[str] = []
    header = rows[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")

    for row in rows[1:]:
        # Pad or trim to match header length
        cells = row + [""] * (len(header) - len(row))
        cells = cells[: len(header)]
        if not any(c.strip() for c in cells):
            continue
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines), {"row_count": len(rows) - 1}


def _parse_docling_fast(path: Path) -> tuple[str, dict] | None:
    """Try docling without OCR for editable PDFs. Returns None on failure.

    Exports to markdown preserving headings, tables and document structure.
    """
    try:
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.document_converter import DocumentConverter, PdfFormatOption

        pdf_opts = PdfPipelineOptions()
        pdf_opts.do_ocr = False
        pdf_opts.do_table_structure = True

        converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts),
            }
        )
        result = converter.convert(str(path))
        text = result.document.export_to_markdown(
            page_break_placeholder="<!-- PAGE_BREAK -->",
        )
        if len(text.strip()) < 200:
            return None
        page_count = len(result.pages) if hasattr(result, "pages") else None
        meta: dict = {"parser": "docling"}
        if page_count is not None:
            meta["page_count"] = page_count
        return text, meta
    except Exception as exc:
        log_event(logger, 30, "docling fast parse failed", path=str(path), error=str(exc))
        return None


def _has_line_structure(text: str) -> bool:
    """Check if extracted text preserves structural markers at line starts.

    Returns True if articles/sections appear at the beginning of lines,
    which is required for legal chunking regexes to work.
    """
    normalized = _normalize_structural_headers(text)
    art_count = len(_ARTIGO_RE.findall(normalized))
    section_count = len(re.findall(r"^CAP[ÍI]TULO\s", text, re.IGNORECASE | re.MULTILINE))
    return (art_count + section_count) >= 3


def _parse_pdf(path: Path) -> tuple[str, dict]:
    """Parse PDF with best available parser.

    Strategy:
    1. Try docling without OCR (markdown with headings)
    2. Verify structural integrity (articles/sections at line starts)
    3. If structure is broken -> PyMuPDF fallback (preserves line breaks)
    4. If PyMuPDF extracts too little -> docling with OCR (scanned PDF)
    """
    # 1. Docling sem OCR — tenta markdown com headings
    docling_result = _parse_docling_fast(path)
    if docling_result is not None:
        text, meta = docling_result
        if _has_line_structure(text):
            log_event(logger, 20, "PDF parsed with docling (markdown with headings)", path=str(path))
            return docling_result
        log_event(logger, 20, "Docling output lacks line structure, trying PyMuPDF",
                  path=str(path))

    # 2. PyMuPDF (preserva quebras de linha naturais do PDF)
    try:
        import fitz  # PyMuPDF
    except ImportError:
        # Se docling retornou algo, usar mesmo sem estrutura ideal
        if docling_result is not None:
            return docling_result
        log_event(logger, 30, "PyMuPDF not available, falling back to docling+OCR", path=str(path))
        return _parse_docling(path)

    doc = fitz.open(str(path))
    page_count = len(doc)
    pages: list[str] = []

    for page in doc:
        text = page.get_text("text")
        if text.strip():
            pages.append(text.strip())

    doc.close()

    total_chars = sum(len(p) for p in pages)
    if total_chars > 200:
        log_event(logger, 20, "PDF parsed with PyMuPDF fallback (plain text)", path=str(path))
        text = _PAGE_BREAK_MARKER.join(pages)
        return text, {"page_count": page_count, "parser": "pymupdf"}

    # 3. Ultimo recurso: docling com OCR (PDF escaneado)
    log_event(logger, 20, "Both docling-fast and PyMuPDF extracted too little, using docling+OCR",
              path=str(path), pymupdf_chars=total_chars)
    return _parse_docling(path)


def _parse_docling(path: Path) -> tuple[str, dict]:
    """Parse PDF/DOCX/PPTX using docling with OCR enabled."""
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pdf_opts = PdfPipelineOptions()
    pdf_opts.do_ocr = True
    pdf_opts.do_table_structure = True

    converter = DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts),
        }
    )
    result = converter.convert(str(path))
    text = result.document.export_to_markdown(
        page_break_placeholder="<!-- PAGE_BREAK -->",
    )
    page_count = len(result.pages) if hasattr(result, "pages") else None
    meta: dict = {"parser": "docling"}
    if page_count is not None:
        meta["page_count"] = page_count
    return text, meta


def _parse(source: str) -> tuple[str, dict]:
    """Extract raw text from a file. Returns (text, metadata).

    Metadata may include ``page_count`` for PDF/DOCX files.
    For PDF/DOCX, page break markers are embedded in the text so that
    ``_assign_page_numbers`` can map chunks to their source pages.
    """
    path = Path(source)
    suffix = path.suffix.lower()

    if suffix in {".txt", ".md"}:
        for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                return path.read_text(encoding=encoding), {}
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Unable to decode text file: {suffix}")

    if suffix in {".xlsx", ".xls"}:
        return _parse_xlsx(path)

    if suffix == ".csv":
        return _parse_csv(path)

    if suffix == ".pdf":
        return _parse_pdf(path)

    if suffix in {".docx", ".pptx"}:
        return _parse_docling(path)

    raise ValueError(f"Unsupported file type: {suffix}")


def _assign_page_numbers(chunks: list[str], raw_text: str) -> list[int | None]:
    """Map each chunk to its starting page number based on page break markers.

    Returns a list of page numbers (1-indexed) or None if no markers are present.
    """
    # Find positions of all page break markers in raw text
    marker_positions: list[int] = [m.start() for m in _PAGE_BREAK_RE.finditer(raw_text)]
    if not marker_positions:
        return [None] * len(chunks)

    # Clean text (without markers) to find chunk positions
    clean_text = _PAGE_BREAK_RE.sub("", raw_text)

    # Build a mapping: character position in clean text -> page number
    # Page 1 starts at position 0; each marker advances the page
    page_boundaries: list[int] = []  # clean-text positions where a new page starts
    offset_adjustment = 0
    for pos in marker_positions:
        clean_pos = pos - offset_adjustment
        page_boundaries.append(clean_pos)
        offset_adjustment += len("<!-- PAGE_BREAK -->")

    def _page_for_position(pos: int) -> int:
        """Return 1-indexed page number for a character position in clean text."""
        page = 1
        for boundary in page_boundaries:
            if pos >= boundary:
                page += 1
            else:
                break
        return page

    result: list[int | None] = []
    for chunk in chunks:
        idx = clean_text.find(chunk[:80])  # match first 80 chars for robustness
        if idx >= 0:
            result.append(_page_for_position(idx))
        else:
            result.append(None)
    return result


def _split_sentences(text: str) -> list[str]:
    """Split Portuguese text into sentences, respecting common abbreviations."""
    # Protect abbreviations by replacing their dots with a placeholder
    protected = _ABBREVIATIONS.sub(lambda m: m.group(0).replace(".", _ABBREV_PLACEHOLDER), text)
    # Split on sentence-ending punctuation followed by whitespace/newline
    raw_sentences = re.split(r"(?<=[.!?;])\s+", protected)
    # Restore dots in abbreviations
    return [s.replace(_ABBREV_PLACEHOLDER, ".").strip() for s in raw_sentences if s.strip()]


def _split(text: str, chunk_size: int = 512, chunk_overlap: int = 64) -> list[str]:
    """Split text into overlapping chunks respecting sentence boundaries.

    Sentences are greedily packed up to *chunk_size* characters.  Overlap is
    achieved by keeping trailing sentences from the previous chunk that fit
    within *chunk_overlap* characters.  If a single sentence exceeds
    *chunk_size*, it falls back to a character-level split.
    """
    sentences = _split_sentences(text)
    if not sentences:
        return []

    chunks: list[str] = []
    current: list[str] = []
    current_len = 0

    for sent in sentences:
        sent_len = len(sent)

        # Sentence longer than chunk_size: flush current, then char-split
        if sent_len > chunk_size:
            if current:
                chunks.append(" ".join(current))
                current = []
                current_len = 0
            start = 0
            while start < sent_len:
                chunks.append(sent[start : start + chunk_size].strip())
                start += chunk_size - chunk_overlap
            continue

        # Adding this sentence would exceed chunk_size
        if current_len + sent_len + (1 if current else 0) > chunk_size:
            chunks.append(" ".join(current))
            # Overlap: keep trailing sentences that fit within chunk_overlap
            overlap: list[str] = []
            overlap_len = 0
            for s in reversed(current):
                if overlap_len + len(s) + (1 if overlap else 0) > chunk_overlap:
                    break
                overlap.insert(0, s)
                overlap_len += len(s) + (1 if len(overlap) > 1 else 0)
            current = overlap
            current_len = sum(len(s) for s in current) + max(len(current) - 1, 0)

        current.append(sent)
        current_len += sent_len + (1 if len(current) > 1 else 0)

    if current:
        chunks.append(" ".join(current))

    return [c for c in chunks if c]


class IngestionError(Exception):
    """Raised when ingestion fails due to quality gate or other pipeline issues."""
    def __init__(self, message: str, status: str = "failed_parse", quality_score: float = 0.0):
        super().__init__(message)
        self.status = status
        self.quality_score = quality_score


def _ingest_pdf_pipeline(
    source: str,
    doc_id: str,
    settings,
) -> tuple[str, str, dict, list]:
    """Run the PDF pipeline and return (clean_text, raw_text, doc_meta).

    Raises IngestionError if the quality gate rejects the document.
    """
    from src.pdf_pipeline import process_pdf

    result = process_pdf(
        pdf_path=source,
        doc_id=doc_id,
        save_artifacts_flag=settings.pdf_pipeline_save_artifacts,
        artifacts_dir=settings.pdf_pipeline_artifacts_dir,
        min_quality_score=settings.pdf_pipeline_min_quality,
    )

    if result.status == "failed_parse":
        raise IngestionError(
            f"PDF quality too low ({result.quality.score:.2f}): {result.quality.flags}",
            status="failed_parse",
            quality_score=result.quality.score,
        )

    # Use clean text from pipeline (already has PAGE_BREAK markers)
    raw_text = result.clean_text
    doc_meta = {
        "parser": "pdf_pipeline",
        "quality_score": result.quality.score,
        "quality_flags": ",".join(result.quality.flags),
        "pipeline_status": result.status,
        "total_blocks": result.quality.total_blocks,
    }
    if result.md_path:
        doc_meta["md_artifact"] = result.md_path
    if result.json_path:
        doc_meta["json_artifact"] = result.json_path

    return result.clean_text, raw_text, doc_meta, result.blocks


def ingest(
    collection: str,
    source: str,
    doc_id: str | None = None,
    embedding_model: str | None = None,
    workspace_id: str = "default",
    domain_profile: str | None = None,
    stage_callback: Callable[[str, dict], None] | None = None,
) -> int:
    """Parse, chunk, embed and index a document. Returns number of chunks indexed.

    For PDF files with pdf_pipeline_enabled, uses structured block extraction
    with quality gate. Raises IngestionError if quality is below threshold.
    """
    doc_id = doc_id or os.path.basename(source)

    logger.info(f"Ingesting '{source}' into collection '{collection}'")

    from src import llm, vectordb  # lazy imports
    from src.config import get_settings

    settings = get_settings()
    model_name = embedding_model or settings.embedding_model
    physical_collection = vectordb.collection_key(collection, model_name, workspace_id=workspace_id)

    from src.observability import metrics

    # ── PDF pipeline path ────────────────────────────────────────────────
    suffix = Path(source).suffix.lower()
    use_pdf_pipeline = (
        suffix == ".pdf"
        and settings.pdf_pipeline_enabled
    )

    pdf_blocks: list | None = None
    if use_pdf_pipeline:
        with metrics.time_block("ingestion.parse"):
            clean_text, raw_text, doc_meta, pdf_blocks = _ingest_pdf_pipeline(source, doc_id, settings)
    else:
        with metrics.time_block("ingestion.parse"):
            raw_text, doc_meta = _parse(source)
        # Remove page break markers before chunking (they are only used for page mapping)
        clean_text = _PAGE_BREAK_RE.sub("", raw_text)
        # Strip repeated page headers/footers/watermarks that pollute chunks
        clean_text = _strip_page_headers(clean_text)

    # Heavy structural cleanup: merge broken lines, dedup, normalize whitespace
    clean_text = deep_clean_text(clean_text)

    if stage_callback:
        stage_callback("silver_extracted", {"doc_id": doc_id, **doc_meta})

    # Normalize markdown headings (## CAPÍTULO → CAPÍTULO) so regexes work
    clean_text = _normalize_markdown(clean_text)
    clean_text = _normalize_structural_headers(clean_text)
    # Detect document structure (chapters, sections) before chunking
    section_markers = _detect_sections(clean_text)

    from src.classifier import classify_document

    classification = classify_document(
        text=clean_text,
        blocks=pdf_blocks,
        pdf_path=source if suffix == ".pdf" else None,
    )

    forced_type = (domain_profile or "").strip().lower()
    extension_forced_type = ""
    if suffix in {".csv", ".xlsx", ".xls"}:
        extension_forced_type = "tabular"

    if forced_type == "legal":
        doc_type = "legal"
    elif forced_type in {"tabular", "mixed", "narrative"}:
        doc_type = forced_type
    elif extension_forced_type:
        doc_type = extension_forced_type
    else:
        doc_type = classification.doc_type

    if doc_type == "legal":
        chunks, texts_for_embedding, metadatas = _ingest_legal(
            clean_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )
    elif doc_type == "tabular":
        chunks, texts_for_embedding, metadatas = _ingest_tabular(
            clean_text, raw_text, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )
    elif doc_type == "mixed":
        chunks, texts_for_embedding, metadatas = _ingest_mixed(
            clean_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )
    else:
        chunks, texts_for_embedding, metadatas = _ingest_general(
            clean_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )

    # Propagate pipeline metadata to all chunk metadatas
    if use_pdf_pipeline:
        for meta in metadatas:
            meta["parser"] = doc_meta.get("parser", "")
            meta["quality_score"] = doc_meta.get("quality_score", 0.0)
            meta["pipeline_status"] = doc_meta.get("pipeline_status", "")
        texts_for_embedding = _enrich_with_pdf_json_context(texts_for_embedding, metadatas, pdf_blocks or [])

    log_event(
        logger, 20, "Document parsed and chunked",
        collection=collection, embedding_model=model_name, doc_id=doc_id,
        raw_text_length=len(raw_text), chunks=len(chunks),
        sections_detected=len(section_markers), legal=(doc_type == "legal"),
        doc_type=doc_type, table_ratio=classification.table_ratio,
        pdf_pipeline=use_pdf_pipeline,
    )

    ids = [chunk_id(collection, doc_id, i) for i in range(len(chunks))]
    # Update chunk_id in metadatas
    for i, meta in enumerate(metadatas):
        meta["chunk_id"] = ids[i]
        meta["chunk_index"] = i

    if stage_callback:
        stage_callback("gold_indexing", {"doc_id": doc_id, "chunks": len(chunks)})

    with metrics.time_block("ingestion.embed"):
        embeddings = llm.embed(texts_for_embedding, model_name=model_name)

    with metrics.time_block("ingestion.delete_old_chunks"):
        deleted = vectordb.delete_by_doc_id(physical_collection, doc_id)
        if deleted:
            log_event(logger, 20, "Old chunks deleted before re-ingestion",
                      collection=collection, doc_id=doc_id, deleted_chunks=deleted)

    with metrics.time_block("ingestion.vectordb_upsert"):
        vectordb.upsert(physical_collection, ids, embeddings, chunks, metadatas)

    log_event(
        logger, 20, "Document indexed",
        collection=collection, physical_collection=physical_collection,
        embedding_model=model_name, doc_id=doc_id, chunks=len(chunks),
    )

    # ── Legal tree + macro indexing + summaries ────────────────────────
    if doc_type == "legal" and getattr(settings, "legal_tree_enabled", True):
        try:
            _build_and_index_legal_tree(
                clean_text=clean_text,
                doc_id=doc_id,
                collection=collection,
                physical_collection=physical_collection,
                workspace_id=workspace_id,
                model_name=model_name,
                llm_mod=llm,
                vectordb_mod=vectordb,
                metrics=metrics,
                settings=settings,
            )
        except Exception as exc:
            log_event(logger, 30, "Legal tree/summary generation failed (non-fatal)",
                      doc_id=doc_id, error=str(exc))

    if stage_callback:
        stage_callback("indexed", {"doc_id": doc_id, "chunks": len(chunks)})
    return len(chunks)


def _build_and_index_legal_tree(
    clean_text: str,
    doc_id: str,
    collection: str,
    physical_collection: str,
    workspace_id: str,
    model_name: str,
    llm_mod,
    vectordb_mod,
    metrics,
    settings=None,
) -> None:
    """Build legal tree, persist nodes, generate summaries, index macro chunks."""
    from src.legal_tree import build_legal_tree
    from src.summaries import generate_tree_summaries, build_summary_context
    from src import controlplane

    with metrics.time_block("ingestion.legal_tree"):
        tree = build_legal_tree(clean_text, doc_id=doc_id, doc_name=doc_id)

    macro_nodes = tree.get_macro_nodes()
    if not macro_nodes:
        return

    # Persist nodes to SQLite
    controlplane.delete_document_nodes(workspace_id, collection, doc_id)
    for node in macro_nodes:
        controlplane.upsert_document_node(
            workspace_id=workspace_id,
            collection=collection,
            doc_id=doc_id,
            node_id=node.id,
            node_type=node.node_type,
            label=node.label,
            numeral=node.numeral,
            path=node.path,
            parent_node_id=node.parent_id,
            text_length=len(node.text),
            articles=node.articles,
            refs=node.internal_refs,
            tree_json=json.dumps(node.to_dict(), ensure_ascii=False),
        )

    log_event(logger, 20, "Legal tree nodes persisted",
              doc_id=doc_id, nodes=len(macro_nodes))

    # Index macro chunks in FAISS (capítulos/seções as special chunks)
    macro_ids: list[str] = []
    macro_texts: list[str] = []
    macro_embed_texts: list[str] = []
    macro_metadatas: list[dict] = []

    for i, node in enumerate(macro_nodes):
        # Truncate text for embedding but store full context path
        embed_text = f"[{node.node_type.upper()}: {node.label}]\n{node.text[:2000]}"
        store_text = f"[{node.label}]\n{node.text[:4000]}"

        mid = chunk_id(collection, f"{doc_id}::macro", i)
        macro_ids.append(mid)
        macro_texts.append(store_text)
        macro_embed_texts.append(embed_text)
        macro_metadatas.append({
            "doc_id": doc_id,
            "source_filename": doc_id,
            "chunk_index": i,
            "chunk_id": mid,
            "collection": collection,
            "workspace_id": workspace_id,
            "embedding_model": model_name,
            "chunk_type": "macro",
            "node_type": node.node_type,
            "node_id": node.id,
            "label": node.label,
            "numeral": node.numeral,
            "path": node.path,
            "articles_covered": ", ".join(node.articles[:30]),
            "text_length": len(node.text),
        })

    if macro_embed_texts:
        with metrics.time_block("ingestion.embed_macro"):
            macro_embeddings = llm_mod.embed(macro_embed_texts, model_name=model_name)
        with metrics.time_block("ingestion.vectordb_upsert_macro"):
            vectordb_mod.upsert(
                physical_collection, macro_ids, macro_embeddings,
                macro_texts, macro_metadatas,
            )
        log_event(logger, 20, "Macro chunks indexed",
                  doc_id=doc_id, macro_chunks=len(macro_ids))

    # Generate and persist summaries via LLM (controlled by config flags)
    summaries_enabled = getattr(settings, "legal_summaries_enabled", True) if settings else True
    summaries_cache = getattr(settings, "legal_summaries_cache", True) if settings else True

    if not summaries_enabled:
        log_event(logger, 20, "Summaries skipped (legal_summaries_enabled=False)",
                  doc_id=doc_id)
        return

    # Check cache: reuse existing summaries if available
    if summaries_cache:
        existing = controlplane.list_document_summaries(workspace_id, collection, doc_id)
        if existing:
            log_event(logger, 20, "Summaries cache hit — reusing existing summaries",
                      doc_id=doc_id, cached_summaries=len(existing))
            return

    with metrics.time_block("ingestion.generate_summaries"):
        summaries = generate_tree_summaries(tree)

    controlplane.delete_document_summaries(workspace_id, collection, doc_id)
    for summary in summaries:
        controlplane.upsert_document_summary(
            workspace_id=workspace_id,
            collection=collection,
            doc_id=doc_id,
            node_id=summary.node_id,
            node_type=summary.node_type,
            label=summary.label,
            path=summary.path,
            resumo_executivo=summary.resumo_executivo,
            resumo_juridico=summary.resumo_juridico,
            pontos_chave=summary.pontos_chave,
            artigos_cobertos=summary.artigos_cobertos,
            obrigacoes=summary.obrigacoes,
            restricoes=summary.restricoes,
            definicoes=summary.definicoes,
            text_length=summary.text_length,
            source_hash=summary.source_hash,
            source_text_length=summary.source_text_length,
            status=summary.status,
            invalid_reason="; ".join(summary.validation_errors),
            generation_meta=summary.generation_meta,
        )

    log_event(logger, 20, "Document summaries generated and persisted",
              doc_id=doc_id, summaries=len(summaries))


def _ingest_general(
    clean_text, raw_text, section_markers, doc_id, source, collection,
    workspace_id, model_name, settings, metrics,
) -> tuple[list[str], list[str], list[dict]]:
    """General chunking path: fixed-size sentence-aware chunks with section breadcrumbs."""
    with metrics.time_block("ingestion.split"):
        overlap = max(settings.chunk_overlap, getattr(settings, "legal_chunk_overlap", settings.chunk_overlap))
        chunks = _split(clean_text, chunk_size=settings.chunk_size, chunk_overlap=overlap)
    page_numbers = _assign_page_numbers(chunks, raw_text)
    section_info = _assign_sections(chunks, section_markers, clean_text)
    chunks = _prepend_section_context(chunks, section_info)

    source_filename = os.path.basename(source)
    metadatas = [
        {
            "doc_id": doc_id,
            "source": source,
            "source_filename": source_filename,
            "chunk_index": i,
            "chunk_id": "",
            "collection": collection,
            "workspace_id": workspace_id,
            "embedding_model": model_name,
            "chunk_type": "general",
            "adjacent_prev": max(0, i - 1),
            "adjacent_next": min(len(chunks) - 1, i + 1),
            **({"page_number": page_numbers[i]} if page_numbers[i] is not None else {}),
            **section_info[i],
        }
        for i in range(len(chunks))
    ]
    return chunks, chunks, metadatas  # embed same text as stored


def _ingest_legal(
    clean_text, raw_text, section_markers, doc_id, source, collection,
    workspace_id, model_name, settings, metrics,
) -> tuple[list[str], list[str], list[dict]]:
    """Legal chunking path: structural article-based chunks with parent-child hierarchy."""
    with metrics.time_block("ingestion.split_legal"):
        legal_chunks = _split_legal(
            clean_text, section_markers, doc_id,
            max_chunk_size=settings.legal_chunk_max_size,
            child_threshold=settings.legal_child_threshold,
        )

    # Build storage texts and embedding texts separately
    texts_for_storage = [c.text for c in legal_chunks]
    texts_for_embedding = [_build_contextualized_text(c, doc_id) for c in legal_chunks]

    # Assign page numbers to storage texts
    page_numbers = _assign_page_numbers(texts_for_storage, raw_text)

    source_filename = os.path.basename(source)
    metadatas = [
        {
            "doc_id": doc_id,
            "source": source,
            "source_filename": source_filename,
            "chunk_index": i,
            "chunk_id": "",
            "collection": collection,
            "workspace_id": workspace_id,
            "embedding_model": model_name,
            "chunk_type": c.chunk_type,
            "artigo": c.artigo,
            "paragrafo": c.paragrafo,
            "inciso": c.inciso,
            "capitulo": c.capitulo,
            "secao": c.secao,
            "titulo": c.titulo,
            "caminho_hierarquico": c.caminho_hierarquico,
            "parent_key": c.parent_key,
            "references": c.references,
            "adjacent_prev": max(0, i - 1),
            "adjacent_next": min(len(legal_chunks) - 1, i + 1),
            **({"page_number": page_numbers[i]} if page_numbers[i] is not None else {}),
        }
        for i, c in enumerate(legal_chunks)
    ]
    return texts_for_storage, texts_for_embedding, metadatas


def _ingest_tabular(
    clean_text, raw_text, doc_id, source, collection,
    workspace_id, model_name, settings, metrics,
) -> tuple[list[str], list[str], list[dict]]:
    """Tabular ingestion path: structured row extraction + tabular chunks."""
    from src.tabular import chunk_tabular_records, extract_tables_from_text, extract_tables_pdfplumber

    with metrics.time_block("ingestion.split_tabular"):
        if str(source).lower().endswith(".pdf"):
            extraction = extract_tables_pdfplumber(source)
        else:
            extraction = extract_tables_from_text(clean_text)

    # Fallback to general when no tabular rows are extracted.
    if not extraction.records:
        section_markers = _detect_sections(clean_text)
        return _ingest_general(
            clean_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )

    if settings.structured_store_enabled:
        from src.structured_store import persist_table_semantics, upsert_records

        with metrics.time_block("ingestion.structured_store_upsert"):
            try:
                upsert_records(collection, doc_id, extraction.records, extraction.column_names)
                persist_table_semantics(collection, workspace_id=workspace_id, context_hint="")
            except Exception as exc:
                log_event(
                    logger, 30, "Structured store upsert failed; continuing with vector indexing",
                    collection=collection, doc_id=doc_id, error=str(exc),
                )

    chunked = chunk_tabular_records(
        extraction.records,
        group_size=settings.tabular_chunk_group_size,
        max_chunk_chars=settings.tabular_chunk_max_chars,
    )
    source_filename = os.path.basename(source)
    chunks: list[str] = []
    metadatas: list[dict] = []

    for i, (chunk_text, chunk_meta) in enumerate(chunked):
        meta = {
            "doc_id": doc_id,
            "source": source,
            "source_filename": source_filename,
            "chunk_index": i,
            "chunk_id": "",
            "collection": collection,
            "workspace_id": workspace_id,
            "embedding_model": model_name,
            "chunk_type": "tabular",
            "row_start": int(chunk_meta.get("row_start", i)),
            "row_end": int(chunk_meta.get("row_end", i)),
            "row_count": int(chunk_meta.get("row_count", 1)),
            "column_names": ",".join(extraction.column_names),
            "fields_json": json.dumps(chunk_meta.get("fields", {}), ensure_ascii=False),
            "raw_rows": "\n".join(chunk_meta.get("raw_rows", [])),
        }
        if chunk_meta.get("page_number") is not None:
            meta["page_number"] = int(chunk_meta["page_number"])
        fields = chunk_meta.get("fields", {}) or {}
        for key, value in fields.items():
            safe_key = re.sub(r"[^a-zA-Z0-9_]+", "_", str(key)).strip("_").lower()[:40]
            if safe_key:
                meta[f"f_{safe_key}"] = str(value)[:500]

        chunks.append(chunk_text)
        metadatas.append(meta)

    return chunks, chunks, metadatas


def _ingest_mixed(
    clean_text, raw_text, section_markers, doc_id, source, collection,
    workspace_id, model_name, settings, metrics,
) -> tuple[list[str], list[str], list[dict]]:
    """Mixed ingestion path: combine tabular extraction with narrative chunks."""
    tab_chunks, tab_embed, tab_meta = _ingest_tabular(
        clean_text, raw_text, doc_id, source, collection,
        workspace_id, model_name, settings, metrics,
    )

    # Keep prose lines outside markdown tables for narrative retrieval.
    prose_text = "\n".join(line for line in clean_text.splitlines() if "|" not in line).strip()
    gen_chunks: list[str] = []
    gen_embed: list[str] = []
    gen_meta: list[dict] = []
    if prose_text:
        gen_chunks, gen_embed, gen_meta = _ingest_general(
            prose_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )
        for meta in gen_meta:
            meta["chunk_type"] = "mixed_narrative"

    if tab_meta:
        for meta in tab_meta:
            if meta.get("chunk_type") == "tabular":
                meta["chunk_type"] = "mixed_tabular"

    chunks = tab_chunks + gen_chunks
    texts_for_embedding = tab_embed + gen_embed
    metadatas = tab_meta + gen_meta

    # If tabular extraction collapsed to general fallback and prose is empty.
    if not chunks:
        return _ingest_general(
            clean_text, raw_text, section_markers, doc_id, source, collection,
            workspace_id, model_name, settings, metrics,
        )

    # Ensure chunk indices are contiguous before outer ingest assigns ids.
    for idx, meta in enumerate(metadatas):
        meta["chunk_index"] = idx
    return chunks, texts_for_embedding, metadatas
