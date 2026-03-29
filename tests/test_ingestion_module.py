from pathlib import Path

import pytest

from src import ingestion


def test_parse_txt_file(tmp_path: Path):
    path = tmp_path / "doc.txt"
    path.write_text("conteudo simples", encoding="utf-8")
    text, meta = ingestion._parse(str(path))
    assert text == "conteudo simples"
    assert meta == {}


def test_parse_txt_file_with_windows_encoding_fallback(tmp_path: Path):
    path = tmp_path / "doc-cp1252.txt"
    path.write_bytes("ação válida".encode("cp1252"))
    text, meta = ingestion._parse(str(path))
    assert text == "ação válida"


def test_parse_unsupported_suffix(tmp_path: Path):
    path = tmp_path / "doc.xyz"
    path.write_text("data", encoding="utf-8")
    with pytest.raises(ValueError):
        ingestion._parse(str(path))


def test_parse_csv_file(tmp_path: Path):
    path = tmp_path / "data.csv"
    path.write_text("nome,cargo,salario\nJoao,Analista,5000\nMaria,Gerente,8000", encoding="utf-8")
    text, meta = ingestion._parse(str(path))
    assert "Joao" in text
    assert "Analista" in text
    assert "| nome | cargo | salario |" in text
    assert meta["row_count"] == 2


def test_parse_csv_semicolon_delimiter(tmp_path: Path):
    path = tmp_path / "data.csv"
    path.write_text("nome;valor\nItem A;100\nItem B;200", encoding="utf-8")
    text, meta = ingestion._parse(str(path))
    assert "Item A" in text
    assert "100" in text


def test_parse_xlsx_file(tmp_path: Path):
    import openpyxl

    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funcionarios"
    ws.append(["Nome", "Cargo", "Salario"])
    ws.append(["Ana", "Dev", "7000"])
    ws.append(["Pedro", "QA", "6000"])
    wb.save(str(path))

    text, meta = ingestion._parse(str(path))
    assert "Ana" in text
    assert "Dev" in text
    assert "Funcionarios" in text
    assert "| Nome | Cargo | Salario |" in text
    assert meta["sheet_count"] == 1


def test_parse_xlsx_multiple_sheets(tmp_path: Path):
    import openpyxl

    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "RH"
    ws1.append(["Nome", "Depto"])
    ws1.append(["Carlos", "TI"])
    ws2 = wb.create_sheet("Contratos")
    ws2.append(["Fornecedor", "Valor"])
    ws2.append(["ABC Ltda", "50000"])
    wb.save(str(path))

    text, meta = ingestion._parse(str(path))
    assert "Carlos" in text
    assert "ABC Ltda" in text
    assert meta["sheet_count"] == 2


def test_assign_page_numbers_maps_chunks_to_pages():
    raw = "Texto da pagina 1.<!-- PAGE_BREAK -->Texto da pagina 2.<!-- PAGE_BREAK -->Texto da pagina 3."
    chunks = ["Texto da pagina 1.", "Texto da pagina 2.", "Texto da pagina 3."]
    pages = ingestion._assign_page_numbers(chunks, raw)
    assert pages == [1, 2, 3]


def test_assign_page_numbers_returns_none_without_markers():
    raw = "Texto simples sem marcadores de pagina."
    chunks = ["Texto simples sem marcadores de pagina."]
    pages = ingestion._assign_page_numbers(chunks, raw)
    assert pages == [None]


def test_assign_page_numbers_handles_chunk_spanning_pages():
    raw = "Inicio da frase.<!-- PAGE_BREAK -->Continuacao da frase. Mais texto na pagina 2."
    chunks = ["Inicio da frase.", "Continuacao da frase. Mais texto na pagina 2."]
    pages = ingestion._assign_page_numbers(chunks, raw)
    assert pages[0] == 1
    assert pages[1] == 2


def test_ingest_passes_expected_data(monkeypatch, tmp_path: Path):
    path = tmp_path / "doc.txt"
    path.write_text("texto de exemplo para ingestao", encoding="utf-8")

    monkeypatch.setattr("src.ingestion.log_event", lambda *args, **kwargs: None)
    monkeypatch.setattr("src.config.get_settings", lambda: type("S", (), {
        "embedding_model": "model-a",
        "chunk_size": 512,
        "chunk_overlap": 64,
    })())

    captured = {}

    def fake_embed(chunks, model_name=None):
        captured["chunks"] = chunks
        captured["embed_model"] = model_name
        return [[0.1, 0.2] for _ in chunks]

    def fake_upsert(collection_name, ids, embeddings, documents, metadatas):
        captured["collection_name"] = collection_name
        captured["ids"] = ids
        captured["documents"] = documents
        captured["metadatas"] = metadatas

    monkeypatch.setattr("src.llm.embed", fake_embed)
    monkeypatch.setattr("src.vectordb.collection_key", lambda c, m, workspace_id="default": f"{workspace_id}::{c}::{m}")
    monkeypatch.setattr("src.vectordb.delete_by_doc_id", lambda col, doc: 0)
    monkeypatch.setattr("src.vectordb.upsert", fake_upsert)
    monkeypatch.setattr("src.vectordb.list_documents", lambda name: (["id-1"], ["texto"]))
    total = ingestion.ingest("geral", str(path), doc_id="arquivo.txt", embedding_model="model-x")

    assert total >= 1
    assert captured["embed_model"] == "model-x"
    assert captured["collection_name"] == "default::geral::model-x"
    assert captured["metadatas"][0]["doc_id"] == "arquivo.txt"
    assert captured["metadatas"][0]["source_filename"] == "doc.txt"


def test_ingest_deletes_old_chunks_before_upsert(monkeypatch, tmp_path: Path):
    """Verify delete_by_doc_id is called BEFORE upsert to avoid orphaned chunks."""
    path = tmp_path / "doc.txt"
    path.write_text("conteudo para teste de ordem", encoding="utf-8")

    monkeypatch.setattr("src.ingestion.log_event", lambda *args, **kwargs: None)
    monkeypatch.setattr("src.config.get_settings", lambda: type("S", (), {
        "embedding_model": "model-a",
        "chunk_size": 512,
        "chunk_overlap": 64,
    })())

    call_order: list[str] = []

    def fake_delete(col, doc):
        call_order.append("delete")
        return 3

    def fake_upsert(collection_name, ids, embeddings, documents, metadatas):
        call_order.append("upsert")

    monkeypatch.setattr("src.llm.embed", lambda chunks, model_name=None: [[0.1] for _ in chunks])
    monkeypatch.setattr("src.vectordb.collection_key", lambda c, m, workspace_id="default": f"{workspace_id}::{c}::{m}")
    monkeypatch.setattr("src.vectordb.delete_by_doc_id", fake_delete)
    monkeypatch.setattr("src.vectordb.upsert", fake_upsert)
    monkeypatch.setattr("src.vectordb.list_documents", lambda name: (["id-1"], ["texto"]))
    ingestion.ingest("geral", str(path), doc_id="doc.txt", embedding_model="model-x")

    assert call_order == ["delete", "upsert"]


def test_ingest_routes_to_tabular_when_classifier_says_tabular(monkeypatch, tmp_path: Path):
    path = tmp_path / "doc.txt"
    path.write_text("texto tabela", encoding="utf-8")

    monkeypatch.setattr("src.ingestion.log_event", lambda *args, **kwargs: None)
    monkeypatch.setattr("src.config.get_settings", lambda: type("S", (), {
        "embedding_model": "model-a",
        "chunk_size": 512,
        "chunk_overlap": 64,
        "pdf_pipeline_enabled": False,
        "legal_chunk_overlap": 160,
        "tabular_chunk_group_size": 5,
        "tabular_chunk_max_chars": 512,
        "structured_store_enabled": False,
    })())
    monkeypatch.setattr(
        "src.classifier.classify_document",
        lambda **kwargs: type("C", (), {"doc_type": "tabular", "table_ratio": 0.7})(),
    )
    monkeypatch.setattr("src.llm.embed", lambda chunks, model_name=None: [[0.1] for _ in chunks])
    monkeypatch.setattr("src.vectordb.collection_key", lambda c, m, workspace_id="default": f"{workspace_id}::{c}::{m}")
    monkeypatch.setattr("src.vectordb.delete_by_doc_id", lambda *a, **k: 0)
    monkeypatch.setattr("src.vectordb.upsert", lambda *a, **k: None)

    captured = {"tabular_called": False}

    def fake_tabular(*args, **kwargs):
        captured["tabular_called"] = True
        return ["chunk tabular"], ["chunk tabular"], [{"doc_id": "doc.txt", "chunk_type": "tabular"}]

    monkeypatch.setattr("src.ingestion._ingest_tabular", fake_tabular)
    monkeypatch.setattr("src.ingestion._ingest_general", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_legal", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_mixed", lambda *a, **k: ([], [], []))

    total = ingestion.ingest("geral", str(path), doc_id="doc.txt", embedding_model="model-x")
    assert total == 1
    assert captured["tabular_called"] is True


def test_ingest_routes_csv_to_tabular_even_if_classifier_says_narrative(monkeypatch, tmp_path: Path):
    path = tmp_path / "base.csv"
    path.write_text("id,nome,renda\n1,Ana,1000\n2,Bia,2000\n", encoding="utf-8")

    monkeypatch.setattr("src.ingestion.log_event", lambda *args, **kwargs: None)
    monkeypatch.setattr("src.config.get_settings", lambda: type("S", (), {
        "embedding_model": "model-a",
        "chunk_size": 512,
        "chunk_overlap": 64,
        "pdf_pipeline_enabled": False,
        "legal_chunk_overlap": 160,
        "tabular_chunk_group_size": 5,
        "tabular_chunk_max_chars": 512,
        "structured_store_enabled": False,
    })())
    monkeypatch.setattr(
        "src.classifier.classify_document",
        lambda **kwargs: type("C", (), {"doc_type": "narrative", "table_ratio": 0.0})(),
    )
    monkeypatch.setattr("src.llm.embed", lambda chunks, model_name=None: [[0.1] for _ in chunks])
    monkeypatch.setattr("src.vectordb.collection_key", lambda c, m, workspace_id="default": f"{workspace_id}::{c}::{m}")
    monkeypatch.setattr("src.vectordb.delete_by_doc_id", lambda *a, **k: 0)
    monkeypatch.setattr("src.vectordb.upsert", lambda *a, **k: None)

    captured = {"tabular_called": False}

    def fake_tabular(*args, **kwargs):
        captured["tabular_called"] = True
        return ["chunk tabular"], ["chunk tabular"], [{"doc_id": "base.csv", "chunk_type": "tabular"}]

    monkeypatch.setattr("src.ingestion._ingest_tabular", fake_tabular)
    monkeypatch.setattr("src.ingestion._ingest_general", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_legal", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_mixed", lambda *a, **k: ([], [], []))

    total = ingestion.ingest("geral", str(path), doc_id="base.csv", embedding_model="model-x")
    assert total == 1
    assert captured["tabular_called"] is True


def test_ingest_routes_to_mixed_when_classifier_says_mixed(monkeypatch, tmp_path: Path):
    path = tmp_path / "doc.txt"
    path.write_text("texto misto", encoding="utf-8")

    monkeypatch.setattr("src.ingestion.log_event", lambda *args, **kwargs: None)
    monkeypatch.setattr("src.config.get_settings", lambda: type("S", (), {
        "embedding_model": "model-a",
        "chunk_size": 512,
        "chunk_overlap": 64,
        "pdf_pipeline_enabled": False,
        "legal_chunk_overlap": 160,
        "tabular_chunk_group_size": 5,
        "tabular_chunk_max_chars": 512,
        "structured_store_enabled": False,
    })())
    monkeypatch.setattr(
        "src.classifier.classify_document",
        lambda **kwargs: type("C", (), {"doc_type": "mixed", "table_ratio": 0.3})(),
    )
    monkeypatch.setattr("src.llm.embed", lambda chunks, model_name=None: [[0.1] for _ in chunks])
    monkeypatch.setattr("src.vectordb.collection_key", lambda c, m, workspace_id="default": f"{workspace_id}::{c}::{m}")
    monkeypatch.setattr("src.vectordb.delete_by_doc_id", lambda *a, **k: 0)
    monkeypatch.setattr("src.vectordb.upsert", lambda *a, **k: None)

    captured = {"mixed_called": False}

    def fake_mixed(*args, **kwargs):
        captured["mixed_called"] = True
        return ["chunk mixed"], ["chunk mixed"], [{"doc_id": "doc.txt", "chunk_type": "mixed_narrative"}]

    monkeypatch.setattr("src.ingestion._ingest_mixed", fake_mixed)
    monkeypatch.setattr("src.ingestion._ingest_general", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_legal", lambda *a, **k: ([], [], []))
    monkeypatch.setattr("src.ingestion._ingest_tabular", lambda *a, **k: ([], [], []))

    total = ingestion.ingest("geral", str(path), doc_id="doc.txt", embedding_model="model-x")
    assert total == 1
    assert captured["mixed_called"] is True


# ── Section detection tests ─────────────────────────────────────────────────


def test_detect_sections_capitulo():
    text = (
        "Preâmbulo do documento.\n"
        "CAPÍTULO I - DA DENOMINAÇÃO\n"
        "Art. 1 - A sociedade denomina-se...\n"
        "CAPÍTULO II - DA ADMINISTRAÇÃO\n"
        "Art. 10 - A administração será...\n"
    )
    markers = ingestion._detect_sections(text)
    assert len(markers) == 2
    assert markers[0].header == "CAPÍTULO I - DA DENOMINAÇÃO"
    assert markers[0].key == "section"
    assert markers[0].level == 2
    assert markers[1].header == "CAPÍTULO II - DA ADMINISTRAÇÃO"
    assert markers[0].offset < markers[1].offset


def test_detect_sections_mixed_hierarchy():
    text = (
        "TÍTULO I - DISPOSIÇÕES GERAIS\n"
        "CAPÍTULO I - DA DENOMINAÇÃO\n"
        "Art. 1 - Blá blá.\n"
        "SEÇÃO I - DOS DIREITOS\n"
        "Art. 5 - Direitos.\n"
    )
    markers = ingestion._detect_sections(text)
    assert len(markers) == 3
    keys = [m.key for m in markers]
    assert keys == ["title", "section", "subsection"]


def test_detect_sections_empty():
    assert ingestion._detect_sections("") == []
    assert ingestion._detect_sections("Texto simples sem headers.") == []


def test_assign_sections_basic():
    text = (
        "CAPÍTULO I - DA DENOMINAÇÃO\n"
        "A sociedade denomina-se Unimed Porto Alegre.\n"
        "CAPÍTULO II - DA ADMINISTRAÇÃO\n"
        "A administração será exercida por um conselho.\n"
    )
    markers = ingestion._detect_sections(text)
    chunks = [
        "A sociedade denomina-se Unimed Porto Alegre.",
        "A administração será exercida por um conselho.",
    ]
    info = ingestion._assign_sections(chunks, markers, text)
    assert info[0] == {"section": "CAPÍTULO I - DA DENOMINAÇÃO"}
    assert info[1] == {"section": "CAPÍTULO II - DA ADMINISTRAÇÃO"}


def test_prepend_section_context():
    chunks = ["Art. 1 conteudo.", "Art. 10 conteudo.", "Texto sem seção."]
    info = [
        {"title": "TÍTULO I", "section": "CAPÍTULO I"},
        {"section": "CAPÍTULO II"},
        {},
    ]
    result = ingestion._prepend_section_context(chunks, info)
    assert result[0] == "[TÍTULO I > CAPÍTULO I]\nArt. 1 conteudo."
    assert result[1] == "[CAPÍTULO II]\nArt. 10 conteudo."
    assert result[2] == "Texto sem seção."


# --- Legal document detection ---

def test_is_legal_document_true():
    text = (
        "CAPITULO I - DAS DISPOSICOES GERAIS\n"
        "Art. 1. A empresa tem por objeto...\n"
        "Art. 2. O prazo de duracao...\n"
        "SECAO I - DO CAPITAL\n"
        "Art. 3. O capital social e de...\n"
    )
    assert ingestion._is_legal_document(text) is True


def test_is_legal_document_false():
    text = "Este e um relatorio financeiro do trimestre. Receita: R$ 1.000. Despesas: R$ 500."
    assert ingestion._is_legal_document(text) is False


def test_is_legal_document_threshold():
    text = "Art. 1. Foo.\nArt. 2. Bar."
    # Only 2 indicators, default threshold=3
    assert ingestion._is_legal_document(text) is False
    assert ingestion._is_legal_document(text, threshold=2) is True


# --- LegalChunk dataclass ---

def test_legal_chunk_defaults():
    chunk = ingestion.LegalChunk(text="some text", chunk_type="parent")
    assert chunk.artigo == ""
    assert chunk.parent_key == ""
    assert chunk.references == ""


# --- _extract_references ---

def test_extract_references_finds_article_refs():
    text = "Conforme disposto no art. 37 e art. 12 desta lei."
    refs = ingestion._extract_references(text, own_artigo_num="5")
    assert "Art. 12" in refs
    assert "Art. 37" in refs


def test_extract_references_excludes_self():
    text = "Art. 5° Conforme art. 5 e art. 12."
    refs = ingestion._extract_references(text, own_artigo_num="5")
    assert "Art. 5" not in refs
    assert "Art. 12" in refs


def test_extract_references_empty():
    text = "Este é um texto sem referências a artigos."
    refs = ingestion._extract_references(text)
    assert refs == ""


# --- _split_legal ---

def test_split_legal_basic_articles():
    text = (
        "Preambulo do estatuto.\n\n"
        "Art. 1. A empresa tem sede na cidade X.\n\n"
        "Art. 2. O objeto social e Y.\n\n"
        "Art. 3. O capital social e Z.\n"
    )
    chunks = ingestion._split_legal(text, [], "doc1")
    # Preamble + 3 parent articles
    types = [c.chunk_type for c in chunks]
    assert types.count("general") == 1  # preamble
    assert types.count("parent") == 3
    assert chunks[1].artigo == "Art. 1"
    assert chunks[2].artigo == "Art. 2"
    assert chunks[3].artigo == "Art. 3"


def test_split_legal_with_paragraphs():
    # Article long enough to trigger child splitting
    long_art = "Art. 1. Caput do artigo com texto extenso.\n" + "Texto " * 100 + "\n"
    long_art += "§ 1. Primeiro paragrafo do artigo.\n"
    long_art += "§ 2. Segundo paragrafo do artigo.\n"

    chunks = ingestion._split_legal(long_art, [], "doc1", child_threshold=200)
    types = [c.chunk_type for c in chunks]
    assert "parent" in types
    assert "child" in types
    # Should have: parent + caput child + §1 child + §2 child
    child_chunks = [c for c in chunks if c.chunk_type == "child"]
    assert len(child_chunks) >= 2


def test_split_legal_parent_key():
    text = "Art. 41. E vedada a remuneracao.\n"
    chunks = ingestion._split_legal(text, [], "estatuto")
    parent = [c for c in chunks if c.chunk_type == "parent"][0]
    assert parent.parent_key == "estatuto::Art. 41"


def test_split_legal_preserves_full_parent_text_even_when_long():
    text = "Art. 1. " + ("Texto muito longo. " * 300)
    chunks = ingestion._split_legal(text, [], "estatuto", max_chunk_size=120, child_threshold=500)
    parent = [c for c in chunks if c.chunk_type == "parent"][0]
    assert len(parent.text) > 120
    assert parent.text.startswith("Art. 1.")


def test_split_legal_no_articles():
    text = "Este texto nao contem artigos, apenas texto corrido."
    chunks = ingestion._split_legal(text, [], "doc1")
    assert len(chunks) == 1
    assert chunks[0].chunk_type == "general"


def test_split_legal_references_populated():
    text = "Art. 10. Conforme previsto no art. 5, aplicam-se as regras do art. 8.\n"
    chunks = ingestion._split_legal(text, [], "doc1")
    parent = [c for c in chunks if c.chunk_type == "parent"][0]
    assert "Art. 5" in parent.references
    assert "Art. 8" in parent.references


def test_split_legal_creates_child_chunks_for_incisos():
    text = (
        "Art. 7. Compete ao cooperado:\n"
        "I - cumprir o estatuto social;\n"
        "II - comparecer as assembleias;\n"
        "III - manter cadastro atualizado.\n"
    )
    chunks = ingestion._split_legal(text, [], "doc1", child_threshold=40)
    children = [c for c in chunks if c.chunk_type == "child"]
    assert any(c.inciso == "I" for c in children)
    assert any(c.inciso == "II" for c in children)
    assert any(c.inciso == "III" for c in children)


# --- _build_contextualized_text ---

def test_build_contextualized_text_with_path():
    chunk = ingestion.LegalChunk(
        text="É vedada a remuneração.",
        chunk_type="parent",
        artigo="Art. 41",
        caminho_hierarquico="CAPÍTULO V > Art. 41",
    )
    result = ingestion._build_contextualized_text(chunk, doc_id="estatuto.pdf")
    assert "[Documento: estatuto.pdf]" in result
    assert "[Caminho: CAPÍTULO V > Art. 41]" in result
    assert "É vedada a remuneração." in result


def test_build_contextualized_text_no_doc_id():
    chunk = ingestion.LegalChunk(text="Texto simples.", chunk_type="general")
    result = ingestion._build_contextualized_text(chunk)
    assert "[Documento:" not in result
    assert "Texto simples." in result


def test_normalize_structural_headers_unwraps_bracketed_lines():
    text = "[CAPÃTULO I - DISPOSICOES GERAIS]\n[SEÃ‡ÃƒO II - DIREITOS]\nArt. 1 - Texto."
    normalized = ingestion._normalize_structural_headers(text)
    assert "[CAPÃTULO I" not in normalized
    assert "CAPÃTULO I - DISPOSICOES GERAIS" in normalized


# ── PDF parsing strategy tests ──────────────────────────────────────────


def test_parse_pdf_prefers_docling_when_structure_preserved(monkeypatch):
    """When docling-fast returns text with line structure, use it over PyMuPDF."""
    md_text = "# Titulo\n\nArt. 1 - Primeiro artigo.\nArt. 2 - Segundo.\nArt. 3 - Terceiro.\n" + "Conteudo extra. " * 20
    monkeypatch.setattr(
        ingestion, "_parse_docling_fast",
        lambda path: (md_text, {"parser": "docling"}),
    )

    text, meta = ingestion._parse_pdf(__import__("pathlib").Path("fake.pdf"))
    assert meta["parser"] == "docling"


def test_parse_pdf_rejects_docling_without_line_structure(monkeypatch):
    """When docling output lacks articles at line starts, fall back to PyMuPDF."""
    # Docling returns text with articles inline (not at line starts)
    md_text = "Texto corrido Art. 1 junto Art. 2 junto Art. 3 sem quebra " * 5
    monkeypatch.setattr(
        ingestion, "_parse_docling_fast",
        lambda path: (md_text, {"parser": "docling"}),
    )

    import types
    fake_page = types.SimpleNamespace(get_text=lambda fmt: "Art. 1 - Texto\nArt. 2 - Mais\nArt. 3 - Outro\n" + "Conteudo. " * 30)

    class FakeDoc:
        def __init__(self):
            self._pages = [fake_page]
        def __iter__(self):
            return iter(self._pages)
        def __len__(self):
            return len(self._pages)
        def close(self):
            pass

    fake_fitz = types.SimpleNamespace(open=lambda path: FakeDoc())
    monkeypatch.setitem(__import__("sys").modules, "fitz", fake_fitz)

    text, meta = ingestion._parse_pdf(__import__("pathlib").Path("fake.pdf"))
    assert meta["parser"] == "pymupdf"


def test_parse_pdf_falls_back_to_pymupdf_when_docling_fails(monkeypatch):
    """When docling-fast returns None, PyMuPDF should be used."""
    monkeypatch.setattr(ingestion, "_parse_docling_fast", lambda path: None)

    import types
    fake_page = types.SimpleNamespace(get_text=lambda fmt: "Texto puro " * 50)
    class FakeDoc:
        def __init__(self):
            self._pages = [fake_page]
        def __iter__(self):
            return iter(self._pages)
        def __len__(self):
            return len(self._pages)
        def close(self):
            pass

    fake_fitz = types.SimpleNamespace(open=lambda path: FakeDoc())
    monkeypatch.setitem(__import__("sys").modules, "fitz", fake_fitz)

    text, meta = ingestion._parse_pdf(__import__("pathlib").Path("fake.pdf"))
    assert meta["parser"] == "pymupdf"
    assert "Texto puro" in text


def test_parse_pdf_falls_back_to_ocr_when_both_fail(monkeypatch):
    """When docling-fast and PyMuPDF both extract too little, docling+OCR is used."""
    monkeypatch.setattr(ingestion, "_parse_docling_fast", lambda path: None)

    import types
    fake_page = types.SimpleNamespace(get_text=lambda fmt: "ab")

    class FakeDoc:
        def __init__(self):
            self._pages = [fake_page]
        def __iter__(self):
            return iter(self._pages)
        def __len__(self):
            return len(self._pages)
        def close(self):
            pass
    fake_fitz = types.SimpleNamespace(open=lambda path: FakeDoc())
    monkeypatch.setitem(__import__("sys").modules, "fitz", fake_fitz)

    monkeypatch.setattr(
        ingestion, "_parse_docling",
        lambda path: ("OCR markdown content " * 20, {"parser": "docling"}),
    )

    text, meta = ingestion._parse_pdf(__import__("pathlib").Path("scan.pdf"))
    assert meta["parser"] == "docling"
    assert "OCR markdown" in text


def test_parse_docling_fast_returns_none_on_short_text(monkeypatch):
    """_parse_docling_fast returns None when extracted text is too short."""
    import types

    class FakeResult:
        def __init__(self):
            self.document = types.SimpleNamespace(
                export_to_markdown=lambda **kw: "short"
            )

    class FakeConverter:
        def __init__(self, **kw):
            pass
        def convert(self, path):
            return FakeResult()

    monkeypatch.setattr(
        "src.ingestion._parse_docling_fast.__module__", "src.ingestion",
        raising=False,
    )
    # Directly test the function with mocked docling imports
    from unittest.mock import patch, MagicMock
    mock_input = MagicMock()
    mock_opts = MagicMock()
    mock_converter_cls = MagicMock(return_value=FakeConverter())

    with patch.dict("sys.modules", {
        "docling": MagicMock(),
        "docling.datamodel": MagicMock(),
        "docling.datamodel.base_models": MagicMock(InputFormat=mock_input),
        "docling.datamodel.pipeline_options": MagicMock(PdfPipelineOptions=lambda: mock_opts),
        "docling.document_converter": MagicMock(
            DocumentConverter=FakeConverter,
            PdfFormatOption=lambda **kw: None,
        ),
    }):
        result = ingestion._parse_docling_fast(__import__("pathlib").Path("fake.pdf"))
    assert result is None


# ── Markdown normalization tests ─────────────────────────────────────────


def test_normalize_markdown_strips_heading_prefixes():
    text = "## CAPÍTULO I - Disposições Gerais\n\n### SEÇÃO II - Direitos\n\nTexto normal."
    result = ingestion._normalize_markdown(text)
    assert result == "CAPÍTULO I - Disposições Gerais\n\nSEÇÃO II - Direitos\n\nTexto normal."


def test_normalize_markdown_preserves_non_heading_text():
    text = "Art. 1 - Texto sem headings.\n| col1 | col2 |\nOutra linha."
    result = ingestion._normalize_markdown(text)
    assert result == text


def test_detect_sections_works_after_markdown_normalization():
    """Full pipeline: markdown headings → normalize → detect sections."""
    md_text = "## CAPÍTULO I - Dos Direitos\n\nArt. 1 - Foo.\n\n### SEÇÃO II - Deveres\n\nArt. 2 - Bar."
    normalized = ingestion._normalize_markdown(md_text)
    markers = ingestion._detect_sections(normalized)
    assert len(markers) >= 2
    headers = [m.header for m in markers]
    assert any("CAPÍTULO I" in h for h in headers)
    assert any("SEÇÃO II" in h or "SECÃO II" in h for h in headers)
